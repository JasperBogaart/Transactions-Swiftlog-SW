import os
import sys
import warnings
from collections import Counter
from datetime import date, datetime
import re
from tkinter import Button, Tk, StringVar, filedialog, messagebox
from tkinter import ttk

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.data_source import AxDataSource, StrRef
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "De module 'openpyxl' is niet geinstalleerd. Installeer deze met: pip install openpyxl"
    ) from exc

warnings.filterwarnings("ignore", message="Workbook contains no default style.*")


def get_program_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


PROGRAM_DIR = get_program_dir()
REPORT_DIR = os.path.join(PROGRAM_DIR, "data")
SYSTEM_DIR = os.path.join(PROGRAM_DIR, "SYSTEM")
DEFAULT_SOURCE = os.path.join(PROGRAM_DIR, "S04 - Transactions.xlsx")
GENERATED_REPORT_PREFIX = "Transactie Overzicht "
MANAGEMENT_REPORT_FILENAME = "Management Overzicht Weekanalyse.xlsx"
PRODUCTIVITY_CONFIG_FILENAME = "productiviteiten.xlsx"
PRODUCTIVITY_CONFIG_PATH = os.path.join(SYSTEM_DIR, PRODUCTIVITY_CONFIG_FILENAME)

REQUIRED_COLUMNS = {
    "User": "User",
    "Pallet Id": "Pallet Id",
    "Activity Description": "Activity Description",
}

DEFAULT_MINUTES_PER_ACTIVITY = {
    "Pallet Pick": 3.75, # 16 per uur
    "Trailer load": 2.15,  # 45 min per vracht van 21 pallets
    "Receiving": 2.15, # 45 min per vracht van 21 pallets 
    "Undirected Full Inventory Move with Putaway": 3.33, # 18 per uur
    "Undirected Full Inventory Move": 3.33, # 18 per uur
    "Assumed Pick": 20, # 3 per uur
    "Pallet Replenishment": 4.3, # 14 per uur 
}


def create_productivity_config_file(path=PRODUCTIVITY_CONFIG_PATH, overwrite=False):
    if os.path.exists(path) and not overwrite:
        return path

    output_dir = os.path.dirname(path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Productiviteiten"
    ws.append(["Activiteit", "Minuten per transactie", "Toelichting"])
    for activity, minutes in DEFAULT_MINUTES_PER_ACTIVITY.items():
        ws.append([activity, minutes, "Pas alleen de minuten aan; de activiteitnaam moet gelijk blijven."])

    style_header(ws[1])
    table_border = make_cell_border()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        row[0].number_format = "0.00"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    autosize_columns(ws, max_width=55)
    wb.save(path)
    return path


def load_productivity_minutes(path=PRODUCTIVITY_CONFIG_PATH):
    create_productivity_config_file(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Productiviteiten" not in wb.sheetnames:
            raise ValueError(f"{os.path.basename(path)} mist tabblad 'Productiviteiten'.")

        ws = wb["Productiviteiten"]
        loaded_minutes = {}
        ordered_activities = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            activity = clean_value(row[0] if row else "")
            if not activity:
                continue
            try:
                minutes = float(row[1])
            except (TypeError, ValueError):
                minutes = DEFAULT_MINUTES_PER_ACTIVITY.get(activity, 0)
            if activity not in loaded_minutes:
                ordered_activities.append(activity)
            loaded_minutes[activity] = minutes
    finally:
        wb.close()

    for activity, minutes in DEFAULT_MINUTES_PER_ACTIVITY.items():
        if activity not in loaded_minutes:
            ordered_activities.append(activity)
            loaded_minutes[activity] = minutes

    return ordered_activities, loaded_minutes


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def find_header_map(header_row):
    normalized = {clean_value(value).lower(): index + 1 for index, value in enumerate(header_row)}
    header_map = {}
    missing = []

    for expected in REQUIRED_COLUMNS:
        key = expected.lower()
        if key not in normalized:
            missing.append(expected)
        else:
            header_map[expected] = normalized[key]

    if missing:
        raise ValueError(
            "Deze verplichte kolommen ontbreken in het bronbestand: " + ", ".join(missing)
        )

    return header_map


def read_unique_transactions(source_path):
    wb = load_workbook(source_path, read_only=True, data_only=True)
    ws = wb.active

    try:
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration as exc:
        wb.close()
        raise ValueError("Het bronbestand is leeg.") from exc

    columns = find_header_map(header)
    user_col = columns["User"]
    pallet_col = columns["Pallet Id"]
    activity_col = columns["Activity Description"]
    transaction_date_col = 2

    unique_keys = set()
    unique_rows = []
    source_rows = 0
    skipped_rows = 0

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        source_rows += 1
        user = clean_value(row[user_col - 1] if len(row) >= user_col else "")
        pallet_id = clean_value(row[pallet_col - 1] if len(row) >= pallet_col else "")
        activity = clean_value(row[activity_col - 1] if len(row) >= activity_col else "")
        transaction_datetime = parse_transaction_datetime(
            row[transaction_date_col - 1] if len(row) >= transaction_date_col else None
        )

        if not user or not pallet_id or not activity:
            skipped_rows += 1
            continue

        key = (user.upper(), activity.upper(), pallet_id.upper())
        if key in unique_keys:
            continue

        unique_keys.add(key)
        unique_rows.append(
            {
                "user": user,
                "activity": activity,
                "pallet_id": pallet_id,
                "transaction_datetime": transaction_datetime,
                "source_row": row_number,
            }
        )

    wb.close()
    return unique_rows, source_rows, skipped_rows, ws.title


def is_valid_source_workbook(path):
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False

    try:
        ws = wb.active
        try:
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            return False
        find_header_map(header)
        return True
    except ValueError:
        return False
    finally:
        wb.close()


def find_default_source(program_dir):
    candidates = []
    for filename in os.listdir(program_dir):
        if not filename.lower().endswith(".xlsx"):
            continue
        if filename.startswith("~$") or filename.startswith(GENERATED_REPORT_PREFIX):
            continue
        path = os.path.join(program_dir, filename)
        if os.path.isfile(path):
            candidates.append(path)

    candidates.sort(key=os.path.getmtime, reverse=True)
    for path in candidates:
        if is_valid_source_workbook(path):
            return path
    return ""


def parse_transaction_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            pass
    return None


def get_transaction_datetime(source_path):
    wb = load_workbook(source_path, read_only=True, data_only=True)
    ws = wb.active
    try:
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            transaction_datetime = parse_transaction_datetime(row[0])
            if transaction_datetime is not None:
                return transaction_datetime
    finally:
        wb.close()
    return None


def build_summary(unique_rows):
    by_user_activity = Counter((row["user"], row["activity"]) for row in unique_rows)
    by_user = Counter(row["user"] for row in unique_rows)
    by_activity = Counter(row["activity"] for row in unique_rows)

    user_activity_rows = [
        (user, activity, count)
        for (user, activity), count in by_user_activity.items()
    ]
    user_activity_rows.sort(key=lambda item: (item[0].lower(), -item[2], item[1].lower()))

    user_rows = sorted(by_user.items(), key=lambda item: (-item[1], item[0].lower()))
    activity_rows = sorted(by_activity.items(), key=lambda item: (-item[1], item[0].lower()))
    detail_rows = sorted(
        unique_rows,
        key=lambda item: (item["user"].lower(), item["activity"].lower(), item["pallet_id"].lower()),
    )

    return user_activity_rows, user_rows, activity_rows, detail_rows


def style_header(row_cells):
    fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(bottom=Side(style="thin", color="A6A6A6"))
    for cell in row_cells:
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws, max_width=55):
    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), max_width)


def make_cell_border(color="D9D9D9"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def copy_column_widths(source_ws, target_ws, last_col):
    for col_idx in range(1, last_col + 1):
        column_letter = get_column_letter(col_idx)
        target_ws.column_dimensions[column_letter].width = (
            source_ws.column_dimensions[column_letter].width
        )


def add_table_like_format(ws, header_row, first_data_row, last_data_row, last_col):
    style_header(ws[header_row])
    table_border = make_cell_border("D9E2F3")
    for row in ws.iter_rows(min_row=first_data_row, max_row=last_data_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if last_data_row >= first_data_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_data_row}"
    ws.freeze_panes = f"A{first_data_row}"


def create_report(source_path, output_path):
    unique_rows, source_rows, skipped_rows, _source_sheet = read_unique_transactions(source_path)

    activities, activity_minutes = load_productivity_minutes()
    users = sorted(
        {row["user"] for row in unique_rows if row["activity"] in activities},
        key=str.lower,
    )
    counts = Counter((row["user"], row["activity"]) for row in unique_rows)

    wb = Workbook()
    counts_ws = create_counts_sheet(wb, users, activities, counts)
    productivity_ws = create_productivity_sheet(wb, users, activities, counts, activity_minutes)
    copy_column_widths(productivity_ws, counts_ws, len(activities) + 2)
    create_user_mix_sheet(wb, users, activities, counts, unique_rows)
    create_hourly_transactions_sheet(wb, users, activities, unique_rows)

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    return len(unique_rows), source_rows, skipped_rows


def create_counts_sheet(wb, users, activities, counts):
    ws = wb.active
    ws.title = "Aantallen"

    header_row_1 = ["user"] + [split_activity_header(activity)[0] for activity in activities] + ["Totaal aantal"]
    header_row_2 = [None] + [split_activity_header(activity)[1] for activity in activities] + ["Transacties"]
    ws.append(header_row_1)
    ws.append(header_row_2)

    activity_totals = [sum(counts[(user, activity)] for user in users) for activity in activities]
    ws.append(["TOTAAL                          =>"] + activity_totals + [sum(activity_totals)])

    for user in users:
        row_counts = [counts[(user, activity)] or None for activity in activities]
        total = sum(value or 0 for value in row_counts)
        ws.append([user] + row_counts + [total])

    last_col = len(activities) + 2
    last_row = ws.max_row
    style_header(ws[1])
    style_header(ws[2])

    table_border = make_cell_border()
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[1] + ws[2] + ws[3]:
        cell.font = Font(bold=True)

    total_col = get_column_letter(last_col)
    for cell in ws[total_col]:
        cell.font = Font(bold=True)

    ws.auto_filter.ref = f"A3:{total_col}{last_row}"
    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    autosize_columns(ws, max_width=28)
    return ws


def read_report_counts(report_path, activities=None):
    wb = load_workbook(report_path, read_only=True, data_only=True)
    try:
        if "Aantallen" not in wb.sheetnames:
            raise ValueError(f"{os.path.basename(report_path)} heeft geen tabblad 'Aantallen'.")

        ws = wb["Aantallen"]
        if activities is None:
            activities, _activity_minutes = load_productivity_minutes()
        activity_columns = {}
        expected_headers = {
            split_activity_header(activity): activity
            for activity in activities
        }

        row_1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        row_2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), ())
        for col_idx, (header_1, header_2) in enumerate(zip(row_1, row_2), start=1):
            activity = expected_headers.get((header_1, header_2))
            if activity:
                activity_columns[activity] = col_idx

        if not activity_columns:
            raise ValueError(
                f"{os.path.basename(report_path)} lijkt geen geldig Transactie Overzicht te zijn."
            )

        counts = Counter()
        users = set()
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row:
                continue
            user = clean_value(row[0])
            if not user:
                continue
            users.add(user)
            for activity, col_idx in activity_columns.items():
                value = row[col_idx - 1] if len(row) >= col_idx else None
                if isinstance(value, (int, float)):
                    counts[(user, activity)] += int(value)

        return users, counts
    finally:
        wb.close()


def create_source_files_sheet(wb, report_paths):
    ws = wb.create_sheet("Bronbestanden")
    ws.append(["Nr", "Bestand", "Pad"])
    for index, path in enumerate(report_paths, start=1):
        ws.append([index, os.path.basename(path), path])

    last_row = ws.max_row
    last_col = ws.max_column
    style_header(ws[1])
    table_border = make_cell_border()
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    autosize_columns(ws, max_width=80)
    return ws


def create_merged_report(report_paths, output_path):
    if len(report_paths) < 2:
        raise ValueError("Selecteer minimaal twee Transactie Overzicht bestanden om samen te voegen.")

    activities, activity_minutes = load_productivity_minutes()
    combined_counts = Counter()
    users = set()

    for report_path in report_paths:
        report_users, report_counts = read_report_counts(report_path, activities)
        users.update(report_users)
        combined_counts.update(report_counts)

    users = sorted(users, key=str.lower)

    wb = Workbook()
    counts_ws = create_counts_sheet(wb, users, activities, combined_counts)
    productivity_ws = create_productivity_sheet(wb, users, activities, combined_counts, activity_minutes)
    copy_column_widths(productivity_ws, counts_ws, len(activities) + 2)
    create_source_files_sheet(wb, report_paths)

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    return len(users), sum(combined_counts.values())


def get_report_week_from_filename(path):
    match = re.search(r"Week\s+(\d{4})-W(\d{1,2})", os.path.basename(path), re.IGNORECASE)
    if match:
        return int(match.group(1)), int(match.group(2))

    report_date = get_report_date_from_filename(path)
    if report_date:
        return report_date.isocalendar()[:2]
    return None


def get_week_label(path):
    report_week = get_report_week_from_filename(path)
    if report_week:
        return f"W{report_week[1]:02d}"
    return os.path.splitext(os.path.basename(path))[0]


def get_chart_week_label(path):
    report_week = get_report_week_from_filename(path)
    if report_week:
        return report_week[1]
    return get_week_label(path)


def get_week_sort_key(path):
    report_week = get_report_week_from_filename(path)
    if report_week:
        iso_year, iso_week = report_week
        return date.fromisocalendar(iso_year, iso_week, 1)
    return date.max


def find_management_week_dir():
    candidates = [
        os.path.join(REPORT_DIR, "week"),
        os.path.join(REPORT_DIR, "per week"),
        os.path.join(PROGRAM_DIR, "week"),
        os.path.join(PROGRAM_DIR, "per week"),
    ]
    for candidate in candidates:
        if os.path.isdir(candidate):
            return candidate
    return ""


def find_management_week_files(week_dir):
    if not os.path.isdir(week_dir):
        raise ValueError(f"De map met weekbestanden bestaat niet: {week_dir}")

    report_paths = []
    for filename in os.listdir(week_dir):
        if filename.startswith("~$") or not filename.lower().endswith(".xlsx"):
            continue
        if filename == MANAGEMENT_REPORT_FILENAME:
            continue
        report_paths.append(os.path.join(week_dir, filename))

    return sorted(report_paths, key=lambda path: (get_week_sort_key(path), os.path.basename(path).lower()))


def read_management_week_row(report_path):
    wb = load_workbook(report_path, read_only=True, data_only=True)
    try:
        if "Aantallen" not in wb.sheetnames:
            raise ValueError(f"{os.path.basename(report_path)} heeft geen tabblad 'Aantallen'.")
        if "Productiviteit" not in wb.sheetnames:
            raise ValueError(f"{os.path.basename(report_path)} heeft geen tabblad 'Productiviteit'.")

        counts_ws = wb["Aantallen"]
        productivity_ws = wb["Productiviteit"]
        header_1 = [counts_ws.cell(row=1, column=col).value for col in range(2, 10)]
        header_2 = [counts_ws.cell(row=2, column=col).value for col in range(2, 10)]
        totals = [counts_ws.cell(row=3, column=col).value or 0 for col in range(2, 10)]
        allowed_hours = productivity_ws["I3"].value or 0

        return {
            "week": get_week_label(report_path),
            "chart_week": get_chart_week_label(report_path),
            "path": report_path,
            "header_1": header_1,
            "header_2": header_2,
            "totals": totals,
            "allowed_hours": allowed_hours,
        }
    finally:
        wb.close()


def create_stock_sheet(wb, management_rows):
    ws = wb.create_sheet("Voorraad ontwikkeling")
    ws.append(["Week", "Grafiek week", "Binnengekomen pallets", "Uitgaande pallets", "Verschil", "Cumulatief verschil"])

    cumulative_difference = 0
    for row_data in management_rows:
        totals = row_data["totals"]
        trailer_load = totals[1] if len(totals) > 1 else 0
        receiving = totals[2] if len(totals) > 2 else 0
        difference = receiving - trailer_load
        cumulative_difference += difference
        ws.append([
            row_data["week"],
            str(row_data["chart_week"]),
            receiving,
            trailer_load,
            difference,
            cumulative_difference,
        ])

    last_row = ws.max_row
    last_col = ws.max_column
    style_header(ws[1])
    table_border = make_cell_border()
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")

    if last_row >= 2:
        positive_fill = PatternFill("solid", fgColor="D9EAD3")
        negative_fill = PatternFill("solid", fgColor="F4CCCC")
        ws.conditional_formatting.add(
            f"E2:E{last_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=positive_fill),
        )
        ws.conditional_formatting.add(
            f"E2:E{last_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=negative_fill),
        )

    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    autosize_columns(ws, max_width=28)
    return ws


def create_management_charts_sheet(wb, stock_ws):
    ws = wb.create_sheet("Grafieken")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.print_area = "A1:N42"

    for col_idx in range(1, 15):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    for row_idx in range(1, 43):
        ws.row_dimensions[row_idx].height = 18

    last_row = stock_ws.max_row
    week_categories = AxDataSource(strRef=StrRef(f=f"'{stock_ws.title}'!$B$2:$B${last_row}"))

    stock_chart = LineChart()
    stock_chart.title = "Cumulatief voorraadverschil"
    stock_chart.y_axis.title = "Pallets"
    stock_chart.x_axis.delete = False
    stock_chart.x_axis.axPos = "b"
    stock_chart.x_axis.tickLblPos = "nextTo"
    stock_chart.x_axis.tickLblSkip = 1
    stock_chart.x_axis.tickMarkSkip = 1
    stock_chart.x_axis.majorTickMark = "out"
    stock_chart.y_axis.axPos = "l"
    stock_chart.add_data(Reference(stock_ws, min_col=6, min_row=1, max_row=last_row), titles_from_data=True)
    for series in stock_chart.series:
        series.cat = week_categories
    stock_chart.height = 9
    stock_chart.width = 25
    stock_chart.dataLabels = DataLabelList()
    stock_chart.dataLabels.showVal = True
    stock_chart.dataLabels.showSerName = False
    stock_chart.dataLabels.showCatName = False
    stock_chart.dataLabels.showLegendKey = False
    stock_chart.legend = None
    ws.add_chart(stock_chart, "A1")

    flow_chart = BarChart()
    flow_chart.type = "col"
    flow_chart.title = "Binnengekomen pallets vs uitgaande pallets"
    flow_chart.y_axis.title = "Aantal pallets"
    flow_chart.x_axis.delete = False
    flow_chart.x_axis.axPos = "b"
    flow_chart.x_axis.tickLblPos = "nextTo"
    flow_chart.x_axis.tickLblSkip = 1
    flow_chart.x_axis.tickMarkSkip = 1
    flow_chart.x_axis.majorTickMark = "out"
    flow_chart.y_axis.axPos = "l"
    flow_chart.add_data(Reference(stock_ws, min_col=3, max_col=4, min_row=1, max_row=last_row), titles_from_data=True)
    for series in flow_chart.series:
        series.cat = week_categories
    flow_chart.height = 9
    flow_chart.width = 25
    flow_chart.dataLabels = DataLabelList()
    flow_chart.dataLabels.showVal = True
    flow_chart.dataLabels.showSerName = False
    flow_chart.dataLabels.showCatName = False
    flow_chart.dataLabels.showLegendKey = False
    flow_chart.dataLabels.position = "inEnd"
    ws.add_chart(flow_chart, "A22")

    return ws


def create_management_week_report(week_dir, output_path):
    report_paths = find_management_week_files(week_dir)
    if not report_paths:
        raise ValueError(f"Geen Excel weekbestanden gevonden in: {week_dir}")

    management_rows = []
    skipped_files = []
    for report_path in report_paths:
        try:
            management_rows.append(read_management_week_row(report_path))
        except Exception as exc:
            skipped_files.append(f"{os.path.basename(report_path)}: {exc}")

    if not management_rows:
        raise ValueError("Geen geldige weekbestanden gevonden.")

    header_1 = ["Week"] + management_rows[0]["header_1"] + ["Uren toegestaan"]
    header_2 = [None] + management_rows[0]["header_2"] + [None]

    wb = Workbook()
    ws = wb.active
    ws.title = "Management overzicht"
    ws.append(header_1)
    ws.append(header_2)
    for row_data in management_rows:
        ws.append([row_data["week"]] + row_data["totals"] + [row_data["allowed_hours"]])

    last_row = ws.max_row
    last_col = ws.max_column
    style_header(ws[1])
    style_header(ws[2])
    table_border = make_cell_border("D9E2F3")
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws["J"]:
        cell.number_format = "0.00"
    ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = "B3"
    ws.sheet_view.showGridLines = False
    autosize_columns(ws, max_width=28)

    stock_ws = create_stock_sheet(wb, management_rows)
    create_management_charts_sheet(wb, stock_ws)
    create_source_files_sheet(wb, [row["path"] for row in management_rows])

    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    return len(management_rows), skipped_files


def create_productivity_sheet(wb, users, activities, counts, minutes_by_activity):
    ws = wb.create_sheet("Productiviteit")

    header_row_1 = ["user"] + [split_activity_header(activity)[0] for activity in activities] + ["Totaal"]
    header_row_2 = [None] + [split_activity_header(activity)[1] for activity in activities] + ["uren"]
    ws.append(header_row_1)
    ws.append(header_row_2)

    total_minutes_by_activity = []
    for activity in activities:
        count_total = sum(counts[(user, activity)] for user in users)
        minute_total = count_total * minutes_by_activity.get(activity, 0)
        total_minutes_by_activity.append(minute_total)
    ws.append(["TOTAAL"] + total_minutes_by_activity + [sum(total_minutes_by_activity) / 60])

    for user in users:
        row = [user]
        total_minutes = 0
        for activity in activities:
            activity_count = counts[(user, activity)]
            minutes = activity_count * minutes_by_activity.get(activity, 0)
            total_minutes += minutes
            row.append(minutes or None)
        row.append(total_minutes / 60)
        ws.append(row)

    last_col = ws.max_column
    last_row = ws.max_row
    last_col_letter = get_column_letter(last_col)

    style_header(ws[1])
    style_header(ws[2])
    table_border = make_cell_border()
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[1] + ws[2] + ws[3]:
        cell.font = Font(bold=True)

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=last_col, max_col=last_col):
        row[0].number_format = "0.00"

    ws.auto_filter.ref = f"A3:{last_col_letter}{last_row}"
    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False
    autosize_columns(ws, max_width=24)
    return ws


def create_user_mix_sheet(wb, users, activities, counts, unique_rows):
    ws = wb.create_sheet("Gebruiker Mix")

    ws.append(["Gebruiker", "Eerste transactie", "Laatste transactie"] + activities)

    transactions_by_user = {user: [] for user in users}
    for row in unique_rows:
        user = row["user"]
        if user in transactions_by_user and row["transaction_datetime"] is not None:
            transactions_by_user[user].append(row["transaction_datetime"])

    for user in users:
        transaction_times = transactions_by_user[user]
        total = sum(counts[(user, activity)] for activity in activities)
        percentages = [
            (counts[(user, activity)] / total if total and counts[(user, activity)] else None)
            for activity in activities
        ]
        ws.append(
            [
                user,
                min(transaction_times) if transaction_times else None,
                max(transaction_times) if transaction_times else None,
            ]
            + percentages
        )

    last_col = ws.max_column
    last_row = ws.max_row
    last_col_letter = get_column_letter(last_col)

    style_header(ws[1])
    table_border = make_cell_border()
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = "hh:mm"

    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=4, max_col=last_col):
        for cell in row:
            cell.number_format = "0.0%"

    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    ws.freeze_panes = "D2"
    ws.sheet_view.showGridLines = False
    autosize_columns(ws, max_width=28)
    return ws


def create_hourly_transactions_sheet(wb, users, activities, unique_rows):
    ws = wb.create_sheet("Transacties per uur")

    time_buckets = ["<7:00"] + list(range(7, 23)) + [">22:00"]
    headers = [
        f"{bucket:02d}:00" if isinstance(bucket, int) else bucket
        for bucket in time_buckets
    ]
    ws.append(["Gebruiker"] + headers + ["Totaal"])

    activity_set = set(activities)
    counts_by_user_bucket = Counter()
    for row in unique_rows:
        transaction_datetime = row["transaction_datetime"]
        if row["activity"] not in activity_set or transaction_datetime is None:
            continue
        if transaction_datetime.hour < 7:
            bucket = "<7:00"
        elif transaction_datetime.hour > 22:
            bucket = ">22:00"
        else:
            bucket = transaction_datetime.hour
        counts_by_user_bucket[(row["user"], bucket)] += 1

    for user in users:
        hourly_counts = [counts_by_user_bucket[(user, bucket)] or None for bucket in time_buckets]
        ws.append([user] + hourly_counts + [sum(value or 0 for value in hourly_counts)])

    last_col = ws.max_column
    last_row = ws.max_row
    last_col_letter = get_column_letter(last_col)

    style_header(ws[1])
    table_border = make_cell_border()
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = table_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for cell in ws[last_col_letter]:
        cell.font = Font(bold=True)

    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    autosize_columns(ws, max_width=14)
    return ws


def split_activity_header(activity):
    header_parts = {
        "Pallet Pick": ("Pallet", "Pick"),
        "Trailer load": ("Trailer", "load"),
        "Receiving": ("Receiving", None),
        "Undirected Full Inventory Move with Putaway": ("Undirected Full", "Inventory Move with Putaway"),
        "Undirected Full Inventory Move": ("Undirected Full", "Inventory Move"),
        "Assumed Pick": ("Assumed", "Pick"),
        "Pallet Replenishment": ("Pallet", "Replenishment"),
    }
    if activity in header_parts:
        return header_parts[activity]

    words = activity.split()
    if len(words) <= 1:
        return activity, None
    midpoint = (len(words) + 1) // 2
    return " ".join(words[:midpoint]), " ".join(words[midpoint:])


def build_output_path(base_dir, report_datetime=None):
    report_datetime = report_datetime or datetime.now()
    filename = f"Transactie Overzicht {report_datetime.strftime('%Y-%m-%d')}.xlsx"
    return os.path.join(base_dir, filename)


def build_default_output_path(source_path=""):
    report_datetime = None
    if source_path:
        try:
            report_datetime = get_transaction_datetime(source_path)
        except Exception:
            report_datetime = None
    return build_output_path(REPORT_DIR, report_datetime)


def get_report_date_from_filename(path):
    match = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path))
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y-%m-%d").date()
    except ValueError:
        return None


def build_merged_output_filename(report_paths=None):
    report_dates = sorted(
        report_date
        for report_date in (get_report_date_from_filename(path) for path in (report_paths or []))
        if report_date is not None
    )

    if report_dates:
        first_date = report_dates[0]
        last_date = report_dates[-1]
        iso_weeks = {report_date.isocalendar()[:2] for report_date in report_dates}
        months = {(report_date.year, report_date.month) for report_date in report_dates}

        if len(iso_weeks) == 1:
            iso_year, iso_week = first_date.isocalendar()[:2]
            return f"Transactie Overzicht Week {iso_year}-W{iso_week:02d}.xlsx"

        if len(months) == 1:
            return f"Transactie Overzicht Maand {first_date.strftime('%Y-%m')}.xlsx"

        return (
            "Transactie Overzicht Periode "
            f"{first_date.strftime('%Y-%m-%d')} tm {last_date.strftime('%Y-%m-%d')}.xlsx"
        )

    return f"Transactie Overzicht Samengevoegd {datetime.now().strftime('%Y-%m-%d')}.xlsx"


def build_merged_output_path(base_dir=REPORT_DIR, report_paths=None):
    filename = build_merged_output_filename(report_paths)
    return os.path.join(base_dir, filename)


def open_output_file(path):
    try:
        os.startfile(path)
    except OSError as exc:
        raise OSError(f"Het rapport is gemaakt, maar kon niet automatisch worden geopend: {exc}") from exc


class TransactionsApp:
    def __init__(self, root):
        self.root = root
        self.root.title("SwiftLog Segway Transactie Overzicht")
        self.root.geometry("760x350")
        self.root.minsize(680, 330)

        default_source = find_default_source(PROGRAM_DIR)
        default_output = build_default_output_path(default_source)

        self.source_var = StringVar(value=default_source)
        self.output_var = StringVar(value=default_output)
        self.status_var = StringVar(value="")
        self.status_label = None

        self.build_ui()
        if default_source:
            self.show_loaded_source(default_source)

    def build_ui(self):
        frame = ttk.Frame(self.root, padding=18)
        frame.pack(fill="both", expand=True)

        frame.columnconfigure(1, weight=1)

        ttk.Label(frame, text="Bronbestand (.xlsx)").grid(row=0, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(frame, textvariable=self.source_var).grid(row=0, column=1, sticky="ew", padx=8, pady=(0, 6))
        ttk.Button(frame, text="Bladeren", command=self.choose_source).grid(row=0, column=2, pady=(0, 6))

        ttk.Label(frame, text="Nieuw rapport").grid(row=1, column=0, sticky="w", pady=(0, 6))
        ttk.Entry(frame, textvariable=self.output_var).grid(row=1, column=1, sticky="ew", padx=8, pady=(0, 6))
        ttk.Button(frame, text="Opslaan als", command=self.choose_output).grid(row=1, column=2, pady=(0, 6))

        ttk.Separator(frame).grid(row=2, column=0, columnspan=3, sticky="ew", pady=14)

        Button(
            frame,
            text="Maak overzicht",
            command=self.run_report,
            bg="#6AA84F",
            activebackground="#93C47D",
            fg="black",
            activeforeground="black",
            font=("Segoe UI", 14, "bold"),
            relief="raised",
            borderwidth=2,
        ).grid(
            row=3, column=0, columnspan=3, sticky="ew", ipady=8
        )
        Button(
            frame,
            text="Maak week/maand overzicht uit meerdere rapporten",
            command=self.run_merged_report,
            bg="#F4CCCC",
            activebackground="#EA9999",
            fg="black",
            activeforeground="black",
            font=("Segoe UI", 14, "bold"),
            relief="raised",
            borderwidth=2,
        ).grid(
            row=4, column=0, columnspan=3, sticky="ew", ipady=8, pady=(8, 0)
        )
        Button(
            frame,
            text="Maak management overzicht uit weekbestanden",
            command=self.run_management_week_report,
            bg="#CFE2F3",
            activebackground="#9FC5E8",
            fg="black",
            activeforeground="black",
            font=("Segoe UI", 14, "bold"),
            relief="raised",
            borderwidth=2,
        ).grid(
            row=5, column=0, columnspan=3, sticky="ew", ipady=8, pady=(8, 0)
        )

        self.status_label = ttk.Label(frame, textvariable=self.status_var, wraplength=700)
        self.status_label.grid(
            row=6, column=0, columnspan=3, sticky="w", pady=(18, 0)
        )

    def set_status(self, text, color="black"):
        self.status_var.set(text)
        if self.status_label is not None:
            self.status_label.configure(foreground=color)

    def show_loaded_source(self, path):
        self.set_status(f"✓ {os.path.basename(path)} geladen", "green")

    def choose_source(self):
        path = filedialog.askopenfilename(
            title="Selecteer WMS transactiebestand",
            filetypes=[("Excel bestanden", "*.xlsx"), ("Alle bestanden", "*.*")],
        )
        if path:
            self.source_var.set(path)
            self.output_var.set(build_default_output_path(path))
            self.show_loaded_source(path)

    def choose_output(self):
        path = filedialog.asksaveasfilename(
            title="Rapport opslaan als",
            defaultextension=".xlsx",
            filetypes=[("Excel bestanden", "*.xlsx")],
            initialfile=os.path.basename(self.output_var.get()) or "Transactie Overzicht.xlsx",
            initialdir=os.path.dirname(self.output_var.get()) or os.getcwd(),
        )
        if path:
            self.output_var.set(path)

    def run_report(self):
        source_path = self.source_var.get().strip()
        output_path = self.output_var.get().strip()

        if not source_path or not os.path.exists(source_path):
            self.set_status("Geen bronbestand geladen.", "red")
            messagebox.showerror("Bronbestand ontbreekt", "Selecteer een bestaand Excel bronbestand.")
            return

        if not output_path:
            self.set_status("Geen rapportpad gekozen.", "red")
            messagebox.showerror("Rapportpad ontbreekt", "Kies waar het nieuwe rapport opgeslagen moet worden.")
            return

        try:
            self.set_status("Bezig met lezen, ontdubbelen en rapport maken...", "black")
            self.root.update_idletasks()
            unique_count, source_rows, skipped_rows = create_report(source_path, output_path)
        except Exception as exc:
            self.set_status("Er is iets misgegaan.", "red")
            messagebox.showerror("Fout bij aanmaken rapport", str(exc))
            return

        self.set_status(
            f"✓ {os.path.basename(source_path)} geladen - klaar: "
            f"{unique_count} unieke transacties uit {source_rows} bronregels.",
            "green",
        )
        try:
            open_output_file(output_path)
        except OSError as exc:
            messagebox.showwarning("Rapport openen mislukt", str(exc))
            return

    def run_merged_report(self):
        report_paths = filedialog.askopenfilenames(
            title="Selecteer Transactie Overzicht bestanden",
            filetypes=[("Excel bestanden", "*.xlsx"), ("Alle bestanden", "*.*")],
            initialdir=REPORT_DIR if os.path.isdir(REPORT_DIR) else PROGRAM_DIR,
        )
        report_paths = [
            path for path in report_paths
            if os.path.basename(path).lower().endswith(".xlsx")
            and not os.path.basename(path).startswith("~$")
        ]
        if not report_paths:
            return

        if len(report_paths) < 2:
            self.set_status("Selecteer minimaal twee rapporten om samen te voegen.", "red")
            messagebox.showerror(
                "Te weinig bestanden",
                "Selecteer minimaal twee Transactie Overzicht bestanden.",
            )
            return

        output_path = filedialog.asksaveasfilename(
            title="Samengevoegd overzicht opslaan als",
            defaultextension=".xlsx",
            filetypes=[("Excel bestanden", "*.xlsx")],
            initialfile=build_merged_output_filename(report_paths),
            initialdir=REPORT_DIR if os.path.isdir(REPORT_DIR) else os.getcwd(),
        )
        if not output_path:
            return

        try:
            self.set_status("Bezig met samenvoegen van bestaande overzichten...", "black")
            self.root.update_idletasks()
            user_count, transaction_count = create_merged_report(report_paths, output_path)
        except Exception as exc:
            self.set_status("Samenvoegen is mislukt.", "red")
            messagebox.showerror("Fout bij samenvoegen", str(exc))
            return

        self.set_status(
            f"Klaar: {len(report_paths)} rapporten samengevoegd tot "
            f"{transaction_count} transacties voor {user_count} gebruikers.",
            "green",
        )
        try:
            open_output_file(output_path)
        except OSError as exc:
            messagebox.showwarning("Rapport openen mislukt", str(exc))
            return

    def run_management_week_report(self):
        week_dir = find_management_week_dir()
        if not week_dir:
            week_dir = filedialog.askdirectory(
                title="Selecteer de map met weekbestanden",
                initialdir=REPORT_DIR if os.path.isdir(REPORT_DIR) else PROGRAM_DIR,
            )
        if not week_dir:
            return

        output_path = filedialog.asksaveasfilename(
            title="Management overzicht opslaan als",
            defaultextension=".xlsx",
            filetypes=[("Excel bestanden", "*.xlsx")],
            initialfile=MANAGEMENT_REPORT_FILENAME,
            initialdir=REPORT_DIR if os.path.isdir(REPORT_DIR) else os.getcwd(),
        )
        if not output_path:
            return

        try:
            self.set_status("Bezig met management overzicht maken uit weekbestanden...", "black")
            self.root.update_idletasks()
            week_count, skipped_files = create_management_week_report(week_dir, output_path)
        except Exception as exc:
            self.set_status("Management overzicht maken is mislukt.", "red")
            messagebox.showerror("Fout bij management overzicht", str(exc))
            return

        self.set_status(f"Klaar: {week_count} weekbestanden verwerkt.", "green")
        if skipped_files:
            messagebox.showwarning(
                "Enkele bestanden overgeslagen",
                "Deze bestanden konden niet worden verwerkt:\n\n" + "\n".join(skipped_files[:10]),
            )
        try:
            open_output_file(output_path)
        except OSError as exc:
            messagebox.showwarning("Rapport openen mislukt", str(exc))
            return


def main():
    if len(sys.argv) >= 4 and sys.argv[1] == "--management-week":
        week_count, skipped_files = create_management_week_report(sys.argv[2], sys.argv[3])
        print(f"Klaar: {week_count} weekbestanden verwerkt.")
        if skipped_files:
            print("Overgeslagen bestanden:")
            for skipped_file in skipped_files:
                print(f"- {skipped_file}")
        return

    if len(sys.argv) >= 4 and sys.argv[1] == "--merge":
        user_count, transaction_count = create_merged_report(sys.argv[3:], sys.argv[2])
        print(
            f"Klaar: {len(sys.argv) - 3} rapporten samengevoegd tot "
            f"{transaction_count} transacties voor {user_count} gebruikers."
        )
        return

    if len(sys.argv) >= 3:
        unique_count, source_rows, skipped_rows = create_report(sys.argv[1], sys.argv[2])
        print(
            f"Klaar: {unique_count} unieke transacties uit {source_rows} bronregels. "
            f"{skipped_rows} regels overgeslagen."
        )
        return

    root = Tk()
    TransactionsApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
